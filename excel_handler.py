"""
Excel handler for FindLEI.

Reads  : .xlsx, .ods (LibreOffice), .xls (legacy)
Detects: LEI column automatically (header name or 20-char alphanumeric pattern)
Writes  : results into two new columns to the right (Entity Status, Next Renewal)
          with colour-coding (green = ACTIVE, orange = other, red = NOT FOUND)
"""

import io
import re
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Constants ────────────────────────────────────────────────────────────────
LEI_RE = re.compile(r"^[A-Z0-9]{18}[0-9]{2}$")
LEI_HEADER_NAMES = {"lei", "lei number", "lei code", "lei_number", "lei_code", "lei no"}

# Column header labels for the two result columns
COL_STATUS  = "Entity Status"
COL_RENEWAL = "Next Renewal Date"
COL_SOURCE  = "Source"

# Palette (ARGB hex, no #)
C_HEADER_STATUS  = "FF1B3A5C"   # dark navy
C_HEADER_RENEWAL = "FF1A6B55"   # dark teal
C_HEADER_SOURCE  = "FF444444"
C_WHITE          = "FFFFFFFF"
C_ACTIVE         = "FFC6EFCE"   # light green
C_ACTIVE_FONT    = "FF1A6B55"
C_INACTIVE       = "FFFFC7CE"   # light red
C_INACTIVE_FONT  = "FF9C0006"
C_UNKNOWN        = "FFFFEB9C"   # light amber
C_UNKNOWN_FONT   = "FF9C5700"


# ── Excel reading ─────────────────────────────────────────────────────────────
class ExcelReadError(Exception):
    pass


def _detect_lei_column(ws) -> Tuple[Optional[int], Optional[int]]:
    """
    Scan the first 15 rows × all columns for a LEI header or a LEI-like value.
    Returns (col_index_1based, header_row_1based).
    header_row is None if the LEI was found in a data cell (no header).
    """
    MAX_SCAN_ROWS = 15

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=MAX_SCAN_ROWS, values_only=True), 1):
        for col_idx, cell_val in enumerate(row, 1):
            if cell_val is None:
                continue
            text = str(cell_val).strip()

            # Header name match
            if text.lower() in LEI_HEADER_NAMES:
                return col_idx, row_idx

            # Data pattern match (20-char alphanumeric ending in 2 digits)
            if LEI_RE.match(text.upper()):
                return col_idx, row_idx - 1  # row above is the "header" row (may be 0)

    return None, None


def read_lei_from_excel(file_bytes: bytes, filename: str) -> Tuple[List[str], Dict]:
    """
    Read LEI numbers from uploaded Excel bytes.

    Returns:
        leis        – ordered list of LEI strings (preserves blank rows as "")
        column_info – metadata dict for write_results_to_excel()
    """
    ext = filename.rsplit(".", 1)[-1].lower()

    if ext in ("xlsx", "xlsm"):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    elif ext == "xls":
        # xlrd path; convert to openpyxl via pandas
        import pandas as pd
        df = pd.read_excel(io.BytesIO(file_bytes), engine="xlrd")
        wb = _df_to_workbook(df)
    elif ext == "ods":
        import pandas as pd
        df = pd.read_excel(io.BytesIO(file_bytes), engine="odf")
        wb = _df_to_workbook(df)
    else:
        raise ExcelReadError(f"Unsupported file type: .{ext}. Use .xlsx, .ods, or .xls")

    ws = wb.active
    lei_col, header_row = _detect_lei_column(ws)

    if lei_col is None:
        raise ExcelReadError(
            "Could not find a LEI column. "
            "Make sure a column is headed 'LEI' or contains 20-character LEI codes."
        )

    data_start_row = (header_row or 0) + 1

    leis: List[str] = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if lei_col - 1 < len(row):
            cell_val = row[lei_col - 1]
            leis.append(str(cell_val).strip() if cell_val is not None else "")
        else:
            leis.append("")

    # Trim trailing blanks
    while leis and not leis[-1]:
        leis.pop()

    column_info = {
        "lei_column":     lei_col,
        "header_row":     header_row,
        "data_start_row": data_start_row,
        "max_row":        ws.max_row,
        "max_col":        ws.max_column,
        "sheet_name":     ws.title,
        "ext":            ext,
    }
    return leis, column_info


def _df_to_workbook(df) -> openpyxl.Workbook:
    """Convert a pandas DataFrame to an openpyxl Workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))
    return wb


# ── Excel writing ─────────────────────────────────────────────────────────────
def write_results_to_excel(
    original_bytes: bytes,
    results: List[Dict],
    column_info: Dict,
) -> bytes:
    """
    Write LEI check results into the original workbook and return modified bytes.

    Results are injected into two (optionally three) columns immediately
    to the right of the last existing column.
    """
    ext = column_info["ext"]

    if ext in ("xlsx", "xlsm"):
        wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    else:
        # For xls/ods we already have an openpyxl Workbook in memory
        # (rebuild from scratch if needed – easier path)
        wb = openpyxl.load_workbook(io.BytesIO(original_bytes))

    ws = wb.active

    lei_col   = column_info["lei_column"]
    hdr_row   = column_info["header_row"]
    start_row = column_info["data_start_row"]
    max_col   = column_info["max_col"]

    status_col  = max_col + 1
    renewal_col = max_col + 2
    source_col  = max_col + 3

    # ── Write column headers ──
    if hdr_row and hdr_row >= 1:
        _write_header(ws, hdr_row, status_col,  COL_STATUS,  C_HEADER_STATUS)
        _write_header(ws, hdr_row, renewal_col, COL_RENEWAL, C_HEADER_RENEWAL)
        _write_header(ws, hdr_row, source_col,  COL_SOURCE,  C_HEADER_SOURCE)

    # ── Index results by LEI ──
    res_map: Dict[str, Dict] = {}
    for r in results:
        lei_key = str(r.get("lei", "")).strip().upper()
        res_map[lei_key] = r

    # ── Write data rows ──
    for row_idx in range(start_row, ws.max_row + 1):
        raw_lei = ws.cell(row=row_idx, column=lei_col).value
        if raw_lei is None:
            continue
        lei_key = str(raw_lei).strip().upper()
        result  = res_map.get(lei_key)
        if not result:
            continue

        status  = result.get("entity_status", "")
        renewal = result.get("next_renewal", "")
        source  = result.get("source", "")

        # Entity Status cell
        sc = ws.cell(row=row_idx, column=status_col, value=status)
        _apply_status_style(sc, status)

        # Next Renewal cell
        rc = ws.cell(row=row_idx, column=renewal_col, value=renewal)
        rc.alignment = Alignment(horizontal="center")

        # Source cell (subtle)
        src_c = ws.cell(row=row_idx, column=source_col, value=source)
        src_c.font = Font(italic=True, color="FF888888", size=9)
        src_c.alignment = Alignment(horizontal="center")

    # ── Auto-fit widths ──
    for col_idx in [status_col, renewal_col, source_col]:
        letter = get_column_letter(col_idx)
        max_len = max(
            (
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range((hdr_row or start_row), ws.max_row + 1)
            ),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max_len + 4, 32)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_header(ws, row: int, col: int, label: str, bg: str):
    cell = ws.cell(row=row, column=col, value=label)
    cell.font      = Font(bold=True, color=C_WHITE, size=10)
    cell.fill      = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_status_style(cell, status: str):
    s = (status or "").upper()
    cell.alignment = Alignment(horizontal="center")
    if "ACTIVE" in s and "INACTIVE" not in s:
        cell.fill = PatternFill(start_color=C_ACTIVE,   end_color=C_ACTIVE,   fill_type="solid")
        cell.font = Font(color=C_ACTIVE_FONT, bold=True)
    elif s in ("NOT FOUND", "INVALID FORMAT") or not s:
        cell.fill = PatternFill(start_color=C_INACTIVE, end_color=C_INACTIVE, fill_type="solid")
        cell.font = Font(color=C_INACTIVE_FONT, bold=True)
    else:
        cell.fill = PatternFill(start_color=C_UNKNOWN,  end_color=C_UNKNOWN,  fill_type="solid")
        cell.font = Font(color=C_UNKNOWN_FONT, bold=True)
