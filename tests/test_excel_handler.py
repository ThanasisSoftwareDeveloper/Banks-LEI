"""Tests for Excel read/write handler."""

import io

import openpyxl
import pytest

from excel_handler import (
    ExcelReadError,
    read_lei_from_excel,
    write_results_to_excel,
    COL_STATUS,
    COL_RENEWAL,
    COL_SOURCE,
)

MOCK_RESULTS = [
    {"lei": "AAAAAA1234567890AA01", "entity_status": "ACTIVE",   "next_renewal": "2025-12-31", "source": "GLEIF"},
    {"lei": "BBBBBB1234567890BB02", "entity_status": "INACTIVE", "next_renewal": "",           "source": "lei-lookup"},
    {"lei": "CCCCCC1234567890CC03", "entity_status": "NOT FOUND","next_renewal": "",           "source": "—"},
    {"lei": "DDDDDD1234567890DD04", "entity_status": "ACTIVE",   "next_renewal": "2026-03-31", "source": "GLEIF"},
]


class TestReadLeiFromExcel:
    def test_detects_lei_column_by_header(self, sample_workbook_bytes):
        leis, info = read_lei_from_excel(sample_workbook_bytes, "test.xlsx")
        assert len(leis) == 4
        assert leis[0] == "AAAAAA1234567890AA01"
        assert info["lei_column"] == 2   # second column
        assert info["header_row"] == 1

    def test_detects_lei_column_by_pattern(self, sample_workbook_no_header_bytes):
        leis, info = read_lei_from_excel(sample_workbook_no_header_bytes, "test.xlsx")
        assert len(leis) >= 1
        assert "AAAAAA1234567890AA01" in leis

    def test_raises_on_unsupported_extension(self, sample_workbook_bytes):
        with pytest.raises(ExcelReadError, match="Unsupported file type"):
            read_lei_from_excel(sample_workbook_bytes, "data.csv")

    def test_raises_when_no_lei_column(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Company", "Country", "Revenue"])
        ws.append(["Acme",    "US",      "1000"])
        buf = io.BytesIO(); wb.save(buf)
        with pytest.raises(ExcelReadError, match="Could not find a LEI column"):
            read_lei_from_excel(buf.getvalue(), "test.xlsx")

    def test_trims_trailing_blank_rows(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["LEI"])
        ws.append(["AAAAAA1234567890AA01"])
        ws.append([None])
        ws.append([None])
        buf = io.BytesIO(); wb.save(buf)
        leis, _ = read_lei_from_excel(buf.getvalue(), "test.xlsx")
        assert leis == ["AAAAAA1234567890AA01"]

    def test_preserves_order(self, sample_workbook_bytes):
        leis, _ = read_lei_from_excel(sample_workbook_bytes, "test.xlsx")
        assert leis[0] == "AAAAAA1234567890AA01"
        assert leis[1] == "BBBBBB1234567890BB02"
        assert leis[2] == "CCCCCC1234567890CC03"
        assert leis[3] == "DDDDDD1234567890DD04"


class TestWriteResultsToExcel:
    def _read_output(self, original_bytes, results, col_info):
        out = write_results_to_excel(original_bytes, results, col_info)
        return openpyxl.load_workbook(io.BytesIO(out)).active

    def test_adds_three_new_columns(self, sample_workbook_bytes):
        leis, info = read_lei_from_excel(sample_workbook_bytes, "test.xlsx")
        ws = self._read_output(sample_workbook_bytes, MOCK_RESULTS, info)
        max_col = info["max_col"]
        assert ws.cell(1, max_col + 1).value == COL_STATUS
        assert ws.cell(1, max_col + 2).value == COL_RENEWAL
        assert ws.cell(1, max_col + 3).value == COL_SOURCE

    def test_active_status_written_correctly(self, sample_workbook_bytes):
        leis, info = read_lei_from_excel(sample_workbook_bytes, "test.xlsx")
        ws = self._read_output(sample_workbook_bytes, MOCK_RESULTS, info)
        status_col = info["max_col"] + 1
        # Row 2 = first data row → ACTIVE
        assert ws.cell(2, status_col).value == "ACTIVE"

    def test_renewal_date_written(self, sample_workbook_bytes):
        leis, info = read_lei_from_excel(sample_workbook_bytes, "test.xlsx")
        ws = self._read_output(sample_workbook_bytes, MOCK_RESULTS, info)
        renewal_col = info["max_col"] + 2
        assert ws.cell(2, renewal_col).value == "2025-12-31"

    def test_not_found_written(self, sample_workbook_bytes):
        leis, info = read_lei_from_excel(sample_workbook_bytes, "test.xlsx")
        ws = self._read_output(sample_workbook_bytes, MOCK_RESULTS, info)
        status_col = info["max_col"] + 1
        # Row 4 = CCCCCC... → NOT FOUND
        assert ws.cell(4, status_col).value == "NOT FOUND"

    def test_output_is_valid_xlsx(self, sample_workbook_bytes):
        leis, info = read_lei_from_excel(sample_workbook_bytes, "test.xlsx")
        out_bytes = write_results_to_excel(sample_workbook_bytes, MOCK_RESULTS, info)
        # openpyxl load without exception means it's valid
        wb = openpyxl.load_workbook(io.BytesIO(out_bytes))
        assert wb is not None
